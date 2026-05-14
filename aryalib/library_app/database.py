import csv
import sqlite3
from contextlib import contextmanager

from library_app.config import (
    BASE_DIR,
    DEFAULT_STUDENTS_FILE,
    EXCEL_STUDENTS_FILE,
    LIBRARY_DATA_FILE,
    LIBRARY_DB_FILE,
    VISITS_FILE,
)
from library_app.time_utils import current_date_text, current_time_text, now_local

_DATABASE_READY = False


@contextmanager
def get_connection():
    conn = sqlite3.connect(LIBRARY_DB_FILE)
    conn.row_factory = sqlite3.Row

    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def _fetchone(conn, query, params=()):
    return conn.execute(query, params).fetchone()


def _fetchall(conn, query, params=()):
    return conn.execute(query, params).fetchall()


def _insert_visit_row(conn, params):
    cursor = conn.execute(
        """
        INSERT INTO visits(student_id, name, father_name, date, entry_time, exit_time)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        params,
    )
    return cursor.lastrowid


def initialize_database():
    with get_connection() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS students (
                student_id TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                father_name TEXT,
                course TEXT,
                phone TEXT,
                valid_until TEXT
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS visits (
                visit_id INTEGER PRIMARY KEY AUTOINCREMENT,
                student_id TEXT NOT NULL,
                name TEXT NOT NULL,
                father_name TEXT,
                date TEXT NOT NULL,
                entry_time TEXT NOT NULL,
                exit_time TEXT DEFAULT '',
                FOREIGN KEY(student_id) REFERENCES students(student_id)
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS sync_state (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS idx_visits_date ON visits(date)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_visits_student_date ON visits(student_id, date)")

        student_columns = {row["name"] for row in conn.execute("PRAGMA table_info(students)").fetchall()}
        if "father_name" not in student_columns:
            conn.execute("ALTER TABLE students ADD COLUMN father_name TEXT DEFAULT ''")

        visit_columns = {row["name"] for row in conn.execute("PRAGMA table_info(visits)").fetchall()}
        if "father_name" not in visit_columns:
            conn.execute("ALTER TABLE visits ADD COLUMN father_name TEXT DEFAULT ''")


def _get_sync_state(conn, key):
    row = _fetchone(conn, "SELECT value FROM sync_state WHERE key = ?", (key,))
    return row["value"] if row else ""


def _set_sync_state(conn, key, value):
    conn.execute(
        """
        INSERT INTO sync_state(key, value)
        VALUES (?, ?)
        ON CONFLICT(key) DO UPDATE SET value = excluded.value
        """,
        (key, value),
    )


def _file_signature(path):
    if not path.exists():
        return ""
    stat = path.stat()
    return f"{path.resolve()}::{stat.st_size}::{stat.st_mtime_ns}"


def _normalize_student_row(row):
    return {
        "student_id": (row.get("student_id") or row.get("Student ID") or "").strip(),
        "name": (row.get("name") or row.get("Name") or "").strip(),
        "father_name": (row.get("father_name") or row.get("Father Name") or row.get("FATHER NAME") or "").strip(),
        "course": (
            row.get("course")
            or row.get("Course")
            or row.get("coursev1")
            or row.get("Coursev1")
            or row.get("branch")
            or row.get("Branch")
            or ""
        ).strip(),
        "phone": (row.get("phone") or row.get("Phone") or "").strip(),
        "valid_until": (row.get("valid_until") or row.get("Valid Until") or "").strip(),
    }


def _student_source_file():
    if EXCEL_STUDENTS_FILE.exists():
        return EXCEL_STUDENTS_FILE
    if LIBRARY_DATA_FILE.exists():
        return LIBRARY_DATA_FILE
    return DEFAULT_STUDENTS_FILE


def _excel_source_files():
    preferred = []
    if EXCEL_STUDENTS_FILE.exists():
        preferred.append(EXCEL_STUDENTS_FILE)

    others = sorted([path for path in BASE_DIR.glob("*.xlsx") if path.name != EXCEL_STUDENTS_FILE.name])
    return preferred + others


def _students_source_signature():
    excel_files = [path for path in _excel_source_files() if path.exists()]
    if excel_files:
        return "|".join(_file_signature(path) for path in excel_files)

    source = _student_source_file()
    return _file_signature(source)


def import_students_from_excel():
    from openpyxl import load_workbook

    excel_files = _excel_source_files()
    if not excel_files:
        return False

    students = {}

    for excel_file in excel_files:
        workbook = load_workbook(excel_file, read_only=True, data_only=True)
        try:
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                rows = worksheet.iter_rows(values_only=True)

                headers = None
                for row in rows:
                    values = [str(cell).strip() if cell is not None else "" for cell in row]
                    if not any(values):
                        continue

                    normalized = [value.lower() for value in values]
                    if headers is None:
                        if (
                            ("name" in normalized or "name " in normalized or "studente name" in normalized or "students name" in normalized)
                            and "branch" in normalized
                            and "code" in normalized
                        ):
                            headers = values
                        continue

                    row_map = dict(zip(headers, values))
                    student_id = str(row_map.get("CODE", "")).strip()
                    name = str(
                        row_map.get("Name ", "")
                        or row_map.get("Name", "")
                        or row_map.get("STUDENTE NAME", "")
                        or row_map.get("Students Name", "")
                    ).strip()
                    course = str(row_map.get("BRANCH", "") or row_map.get("Branch", "")).strip()

                    if not student_id or not name:
                        continue

                    students[student_id] = {
                        "student_id": student_id,
                        "name": name,
                        "father_name": str(row_map.get("FATHER NAME", "") or row_map.get("Father Name", "")).strip(),
                        "course": course,
                        "phone": "",
                        "valid_until": "",
                    }
        finally:
            workbook.close()

    if not students:
        return False

    with get_connection() as conn:
        conn.execute("DELETE FROM students")
        for student in students.values():
            conn.execute(
                """
                INSERT INTO students(student_id, name, father_name, course, phone, valid_until)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    student["student_id"],
                    student["name"],
                    student["father_name"],
                    student["course"],
                    student["phone"],
                    student["valid_until"],
                ),
            )

    return True


def import_students_from_csv():
    source = _student_source_file()
    if source.suffix.lower() == ".xlsx" or not source.exists():
        return False

    students = []
    with source.open("r", newline="", encoding="utf-8-sig") as file:
        reader = csv.DictReader(file)
        for row in reader:
            student = _normalize_student_row(row)
            if not student["student_id"]:
                continue
            students.append(student)

    if not students:
        return False

    with get_connection() as conn:
        conn.execute("DELETE FROM students")
        for student in students:
            conn.execute(
                """
                INSERT INTO students(student_id, name, father_name, course, phone, valid_until)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    student["student_id"],
                    student["name"],
                    student["father_name"],
                    student["course"],
                    student["phone"],
                    student["valid_until"],
                ),
            )

    return True


def import_visits_from_csv():
    if not VISITS_FILE.exists():
        return

    with VISITS_FILE.open("r", newline="", encoding="utf-8") as file, get_connection() as conn:
        count_row = _fetchone(conn, "SELECT COUNT(*) AS total FROM visits")
        if count_row and count_row["total"]:
            return

        for row in csv.DictReader(file):
            student_id = (row.get("student_id") or "").strip()
            name = (row.get("name") or "").strip()
            date = (row.get("date") or row.get("scan_date") or "").strip()
            entry_time = (row.get("entry_time") or "").strip()
            exit_time = (row.get("exit_time") or "").strip()
            father_name = (row.get("father_name") or "").strip()
            if not student_id or not date or not entry_time:
                continue

            _insert_visit_row(conn, (student_id, name, father_name, date, entry_time, exit_time))


def ensure_database_ready():
    global _DATABASE_READY
    if _DATABASE_READY:
        return

    initialize_database()
    student_signature = _students_source_signature()

    with get_connection() as conn:
        current_signature = _get_sync_state(conn, "students_source_signature")
        student_count = _fetchone(conn, "SELECT COUNT(*) AS total FROM students")["total"]
        visits_count = _fetchone(conn, "SELECT COUNT(*) AS total FROM visits")["total"]

    if student_signature and student_count == 0:
        imported = import_students_from_excel() or import_students_from_csv()
        if imported:
            with get_connection() as conn:
                _set_sync_state(conn, "students_source_signature", student_signature)
    elif student_signature and student_count > 0 and current_signature != student_signature:
        with get_connection() as conn:
            _set_sync_state(conn, "students_source_signature", student_signature)

    if visits_count == 0:
        import_visits_from_csv()

    _DATABASE_READY = True


def fetch_students():
    ensure_database_ready()
    with get_connection() as conn:
        rows = _fetchall(
            conn,
            "SELECT student_id, name, father_name, course, phone, valid_until FROM students ORDER BY student_id",
        )
    return {
        row["student_id"]: {
            "student_id": row["student_id"],
            "name": row["name"],
            "father_name": row["father_name"] or "",
            "course": row["course"] or "",
            "phone": row["phone"] or "",
            "valid_until": row["valid_until"] or "",
        }
        for row in rows
    }


def fetch_visits():
    ensure_database_ready()
    with get_connection() as conn:
        rows = _fetchall(
            conn,
            """
            SELECT visit_id, student_id, name, father_name, date, entry_time, exit_time
            FROM visits
            ORDER BY visit_id
            """,
        )
    return [
        {
            "visit_id": str(row["visit_id"]).zfill(5),
            "student_id": row["student_id"],
            "name": row["name"],
            "father_name": row["father_name"] or "",
            "date": row["date"],
            "entry_time": row["entry_time"],
            "exit_time": row["exit_time"] or "",
        }
        for row in rows
    ]


def clear_visits():
    ensure_database_ready()
    with get_connection() as conn:
        conn.execute("DELETE FROM visits")
        conn.execute("DELETE FROM sqlite_sequence WHERE name = ?", ("visits",))


def fetch_latest_visit_for_student(student_id):
    ensure_database_ready()
    with get_connection() as conn:
        row = _fetchone(
            conn,
            """
            SELECT visit_id, student_id, name, father_name, date, entry_time, exit_time
            FROM visits
            WHERE student_id = ?
            ORDER BY date DESC, entry_time DESC, visit_id DESC
            LIMIT 1
            """,
            (student_id,),
        )
    if row is None:
        return None
    return {
        "visit_id": str(row["visit_id"]).zfill(5),
        "student_id": row["student_id"],
        "name": row["name"],
        "father_name": row["father_name"] or "",
        "date": row["date"],
        "entry_time": row["entry_time"],
        "exit_time": row["exit_time"] or "",
    }


def fetch_open_visit(student_id, visit_date):
    ensure_database_ready()
    with get_connection() as conn:
        row = _fetchone(
            conn,
            """
            SELECT visit_id, student_id, name, father_name, date, entry_time, exit_time
            FROM visits
            WHERE student_id = ? AND date = ? AND (exit_time IS NULL OR exit_time = '')
            ORDER BY visit_id DESC
            LIMIT 1
            """,
            (student_id, visit_date),
        )
    if row is None:
        return None
    return {
        "visit_id": str(row["visit_id"]).zfill(5),
        "student_id": row["student_id"],
        "name": row["name"],
        "father_name": row["father_name"] or "",
        "date": row["date"],
        "entry_time": row["entry_time"],
        "exit_time": row["exit_time"] or "",
    }


def create_visit(student):
    now = now_local()
    visit_date = now.date().isoformat()
    entry_time = now.strftime("%H:%M:%S")
    ensure_database_ready()

    with get_connection() as conn:
        visit_id = _insert_visit_row(
            conn,
            (
                student["student_id"],
                student["name"],
                student.get("father_name", ""),
                visit_date,
                entry_time,
                "",
            ),
        )

    return {
        "visit_id": str(visit_id).zfill(5),
        "student_id": student["student_id"],
        "name": student["name"],
        "father_name": student.get("father_name", ""),
        "date": visit_date,
        "entry_time": entry_time,
        "exit_time": "",
    }


def update_visit_exit(student_id, visit_date=None):
    exit_time = current_time_text()
    target_date = visit_date or current_date_text()
    ensure_database_ready()

    with get_connection() as conn:
        row = _fetchone(
            conn,
            """
            SELECT visit_id, student_id, name, father_name, date, entry_time, exit_time
            FROM visits
            WHERE student_id = ? AND date = ? AND (exit_time IS NULL OR exit_time = '')
            ORDER BY visit_id DESC
            LIMIT 1
            """,
            (student_id, target_date),
        )
        if row is None:
            return None

        conn.execute("UPDATE visits SET exit_time = ? WHERE visit_id = ?", (exit_time, row["visit_id"]))

    return {
        "visit_id": str(row["visit_id"]).zfill(5),
        "student_id": row["student_id"],
        "name": row["name"],
        "father_name": row["father_name"] or "",
        "date": row["date"],
        "entry_time": row["entry_time"],
        "exit_time": exit_time,
    }
