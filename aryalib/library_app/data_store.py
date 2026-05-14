import csv
import threading
from datetime import datetime, timedelta

from library_app.config import (
    DEFAULT_STUDENTS_FILE,
    DUPLICATE_SCAN_GAP_SECONDS,
    EXCEL_STUDENTS_FILE,
    LIBRARY_DATA_FILE,
    VISIT_FIELDS,
    VISITS_FILE,
)
from library_app.database import (
    clear_visits,
    create_visit,
    ensure_database_ready,
    fetch_open_visit,
    fetch_latest_visit_for_student,
    fetch_visits,
    update_visit_exit,
)
from library_app.time_utils import current_date_text, now_local, parse_local_timestamp, today_local

_STUDENT_CACHE = {}
_STUDENT_CACHE_SIGNATURE = ""
_STUDENT_CACHE_LOCK = threading.Lock()


def _file_signature(path):
    if not path.exists():
        return ""
    stat = path.stat()
    return f"{path.resolve()}::{stat.st_size}::{stat.st_mtime_ns}"


def _excel_source_files():
    base_dir = EXCEL_STUDENTS_FILE.parent
    preferred = []
    if EXCEL_STUDENTS_FILE.exists():
        preferred.append(EXCEL_STUDENTS_FILE)
    others = sorted([path for path in base_dir.glob("*.xlsx") if path.name != EXCEL_STUDENTS_FILE.name])
    return preferred + others


def _students_source_signature():
    excel_files = [path for path in _excel_source_files() if path.exists()]
    if excel_files:
        return "|".join(_file_signature(path) for path in excel_files)
    return _file_signature(get_students_file())


def _normalize_student_row(row):
    return {
        "student_id": (row.get("student_id") or row.get("Student ID") or "").strip(),
        "name": (row.get("name") or row.get("Name") or "").strip(),
        "father_name": (row.get("father_name") or row.get("Father Name") or row.get("FATHER NAME") or "").strip(),
        "course": (
            row.get("course")
            or row.get("Course")
            or row.get("coursev1")
            or row.get("Coursev1")
            or row.get("branch")
            or row.get("Branch")
            or ""
        ).strip(),
        "phone": (row.get("phone") or row.get("Phone") or "").strip(),
        "valid_until": (row.get("valid_until") or row.get("Valid Until") or "").strip(),
    }


def _load_students_from_excel():
    from openpyxl import load_workbook

    students = {}
    for excel_file in _excel_source_files():
        workbook = load_workbook(excel_file, read_only=True, data_only=True)
        try:
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                rows = worksheet.iter_rows(values_only=True)
                headers = None
                for row in rows:
                    values = [str(cell).strip() if cell is not None else "" for cell in row]
                    if not any(values):
                        continue
                    normalized = [value.lower() for value in values]
                    if headers is None:
                        if (
                            ("name" in normalized or "name " in normalized or "studente name" in normalized or "students name" in normalized)
                            and "branch" in normalized
                            and "code" in normalized
                        ):
                            headers = values
                        continue

                    row_map = dict(zip(headers, values))
                    student_id = str(row_map.get("CODE", "")).strip()
                    name = str(
                        row_map.get("Name ", "")
                        or row_map.get("Name", "")
                        or row_map.get("STUDENTE NAME", "")
                        or row_map.get("Students Name", "")
                    ).strip()
                    if not student_id or not name:
                        continue
                    students[student_id] = {
                        "student_id": student_id,
                        "name": name,
                        "father_name": str(row_map.get("FATHER NAME", "") or row_map.get("Father Name", "")).strip(),
                        "course": str(row_map.get("BRANCH", "") or row_map.get("Branch", "")).strip(),
                        "phone": "",
                        "valid_until": "",
                    }
        finally:
            workbook.close()
    return students


def _load_students_from_csv():
    students = {}
    source = get_students_file()
    if source.suffix.lower() == ".xlsx" or not source.exists():
        return students
    with source.open("r", newline="", encoding="utf-8-sig") as file:
        reader = csv.DictReader(file)
        for row in reader:
            student = _normalize_student_row(row)
            if not student["student_id"]:
                continue
            students[student["student_id"]] = student
    return students


def _load_students_from_source():
    excel_files = [path for path in _excel_source_files() if path.exists()]
    if excel_files:
        return _load_students_from_excel()
    return _load_students_from_csv()


def get_students_file():
    if EXCEL_STUDENTS_FILE.exists():
        return EXCEL_STUDENTS_FILE
    if LIBRARY_DATA_FILE.exists():
        return LIBRARY_DATA_FILE
    return DEFAULT_STUDENTS_FILE


def ensure_students_file():
    ensure_database_ready()


def ensure_visits_file():
    ensure_database_ready()


def load_students():
    global _STUDENT_CACHE
    global _STUDENT_CACHE_SIGNATURE

    signature = _students_source_signature()
    if signature and signature == _STUDENT_CACHE_SIGNATURE and _STUDENT_CACHE:
        return _STUDENT_CACHE

    with _STUDENT_CACHE_LOCK:
        signature = _students_source_signature()
        if signature and signature == _STUDENT_CACHE_SIGNATURE and _STUDENT_CACHE:
            return _STUDENT_CACHE
        _STUDENT_CACHE = _load_students_from_source()
        _STUDENT_CACHE_SIGNATURE = signature
        return _STUDENT_CACHE


def load_visits():
    return fetch_visits()


def save_visits(visits):
    import csv

    with VISITS_FILE.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=VISIT_FIELDS)
        writer.writeheader()
        writer.writerows(visits)
    return visits


def clear_visit_history():
    clear_visits()
    return save_visits([])


def is_membership_valid(student):
    valid_until = student.get("valid_until", "").strip()
    if not valid_until:
        return True, ""

    try:
        expiry_date = datetime.strptime(valid_until, "%Y-%m-%d").date()
    except ValueError:
        return False, "Student date format invalid"

    today = today_local()
    if today > expiry_date:
        return False, f"ID expired on {expiry_date.isoformat()}"

    return True, ""


def find_open_visit(visits, student_id, visit_date):
    for visit in reversed(visits):
        if (
            visit["student_id"] == student_id
            and visit["date"] == visit_date
            and not visit["exit_time"].strip()
        ):
            return visit
    return None


def parse_timestamp(date_text, time_text):
    return parse_local_timestamp(date_text, time_text)


def get_last_scan_timestamp(visits, student_id):
    for visit in reversed(visits):
        if visit["student_id"] != student_id:
            continue
        exit_timestamp = parse_timestamp(visit["date"], visit["exit_time"])
        if exit_timestamp is not None:
            return exit_timestamp
        entry_timestamp = parse_timestamp(visit["date"], visit["entry_time"])
        if entry_timestamp is not None:
            return entry_timestamp
    return None


def process_scan_result(student_id):
    student_id = str(student_id).strip()
    now = now_local()
    today = now.date().isoformat()
    student = load_students().get(student_id)

    if student is None:
        return {
            "ok": False,
            "message": f"Student ID not found: {student_id}",
            "student": None,
            "visit": None,
            "action": "not_found",
        }

    is_valid, reason = is_membership_valid(student)
    if not is_valid:
        return {
            "ok": False,
            "message": reason,
            "student": student,
            "visit": None,
            "action": "invalid",
        }

    last_visit = fetch_latest_visit_for_student(student_id)
    last_scan_timestamp = None
    if last_visit is not None:
        last_scan_timestamp = parse_timestamp(last_visit["date"], last_visit["exit_time"]) or parse_timestamp(
            last_visit["date"], last_visit["entry_time"]
        )
    if last_scan_timestamp is not None:
        if last_scan_timestamp.tzinfo is None:
            last_scan_timestamp = last_scan_timestamp.replace(tzinfo=now.tzinfo)
        elapsed = (now - last_scan_timestamp).total_seconds()
        # A future timestamp can happen if an older record was written with a mismatched
        # server clock or timezone. In that case, don't turn it into a nonsense cooldown.
        if elapsed < 0:
            elapsed = None
        if elapsed is not None and elapsed < DUPLICATE_SCAN_GAP_SECONDS:
            return {
                "ok": False,
                "message": f"Duplicate scan ignored. Try again after {int(DUPLICATE_SCAN_GAP_SECONDS - elapsed) + 1} seconds.",
                "student": student,
                "visit": None,
                "action": "duplicate",
            }

    open_visit = fetch_open_visit(student_id, today)

    if open_visit is None:
        visit = create_visit(student)
        save_visits(load_visits())
        return {
            "ok": True,
            "message": f"Entry saved: {student['name']} ({student['student_id']})",
            "student": student,
            "visit": visit,
            "action": "entry",
        }

    open_visit = update_visit_exit(student_id, today)
    if open_visit is None:
        return {
            "ok": False,
            "message": "Could not update the existing visit. Please try again.",
            "student": student,
            "visit": None,
            "action": "error",
        }
    save_visits(load_visits())
    return {
        "ok": True,
        "message": f"Exit saved: {student['name']} ({student['student_id']})",
        "student": student,
        "visit": open_visit,
        "action": "exit",
    }


def process_scan(student_id):
    result = process_scan_result(student_id)
    return result["ok"], result["message"]


def get_recent_visits(limit=10):
    visits = load_visits()
    recent = list(reversed(visits))
    if limit is not None:
        return recent[:limit]
    return recent


def get_active_visits():
    today = current_date_text()
    return [visit for visit in load_visits() if visit["date"] == today and not visit["exit_time"].strip()]


def get_dashboard_summary():
    students = load_students()
    visits = load_visits()
    today = current_date_text()
    active_visits = [visit for visit in visits if visit["date"] == today and not visit["exit_time"].strip()]
    today_visits = [visit for visit in visits if visit["date"] == today]

    return {
        "student_count": len(students),
        "total_visits": len(visits),
        "today_visits": len(today_visits),
        "inside_count": len(active_visits),
        "today": today,
    }


def with_student_details(visits, students=None):
    student_map = students if students is not None else load_students()
    return [
        {
            **visit,
            "course": student_map.get(visit["student_id"], {}).get("course", visit.get("course", "")),
            "father_name": student_map.get(visit["student_id"], {}).get("father_name", visit.get("father_name", "")),
        }
        for visit in visits
    ]


def build_daily_summary(visits):
    summary_map = {}
    for visit in visits:
        row = summary_map.setdefault(
            visit["date"],
            {"date": visit["date"], "total_visits": 0, "completed_visits": 0, "inside_count": 0},
        )
        row["total_visits"] += 1
        if visit["exit_time"].strip():
            row["completed_visits"] += 1
        else:
            row["inside_count"] += 1
    return sorted(summary_map.values(), key=lambda item: item["date"], reverse=True)


def build_weekly_summary(visits, today=None):
    summary_map = {}
    current_day = today or today_local()
    for visit in visits:
        visit_date = datetime.strptime(visit["date"], "%Y-%m-%d").date()
        iso_year, iso_week, iso_day = visit_date.isocalendar()
        week_key = f"{iso_year}-W{iso_week:02d}"
        start_date = visit_date - timedelta(days=iso_day - 1)
        end_date = start_date + timedelta(days=6)
        row = summary_map.setdefault(
            week_key,
            {
                "week_label": week_key,
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
                "total_visits": 0,
                "completed_visits": 0,
                "inside_count": 0,
                "is_completed": end_date < current_day,
            },
        )
        row["total_visits"] += 1
        if visit["exit_time"].strip():
            row["completed_visits"] += 1
        else:
            row["inside_count"] += 1
    return [
        item
        for item in sorted(summary_map.values(), key=lambda item: item["week_label"], reverse=True)
        if item["is_completed"]
    ]


def build_dashboard_payload(recent_limit=12):
    students = load_students()
    visits = load_visits()
    today = current_date_text()
    recent_visits = [visit for visit in reversed(visits) if visit["date"] == today][:recent_limit]
    active_visits = [visit for visit in visits if visit["date"] == today and not visit["exit_time"].strip()]
    daily_summary = build_daily_summary(visits)
    weekly_summary = build_weekly_summary(visits)

    return {
        "summary": {
            "student_count": len(students),
            "total_visits": len(visits),
            "today_visits": sum(1 for visit in visits if visit["date"] == today),
            "inside_count": len(active_visits),
            "today": today,
        },
        "recent_visits": recent_visits,
        "recent_visits_with_students": with_student_details(recent_visits, students),
        "active_visits": with_student_details(active_visits, students),
        "daily_summary": daily_summary,
        "weekly_summary": weekly_summary,
    }
